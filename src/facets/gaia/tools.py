"""Minimal file tools for the GAIA scenario (the self-contained, no-web slice).

GAIA is fundamentally a web-agent benchmark (~73% of questions need live search), but a small
slice is answerable from a single attached file. This module provides just enough tooling to run
recipes against that slice — enough to prove the scenario abstraction against real GAIA data
*without* a web-search dependency. The full GAIA toolset (web_search / fetch_page, vision, audio)
is deliberately out of scope here; see ``docs/scenarios-gaia.md``.

Tools, scoped to one question's single attachment:

* ``read_file`` — read a text/csv/json/py file (a line window, like OfficeQA's read_document).
* ``read_spreadsheet`` — dump an ``.xlsx`` sheet as rows (lazy ``openpyxl``).
* ``read_pdf`` — extract a PDF's text (lazy ``pypdf``).
* ``compute`` — the same safe arithmetic evaluator OfficeQA uses, reused verbatim.

Binary-format readers (openpyxl, pypdf) are imported lazily and raise an actionable error if
absent, so importing this package never requires them — matching how the dataset client treats
pyarrow.
"""

from __future__ import annotations

from typing import TYPE_CHECKING

from facets.officeqa.tools import _safe_eval  # reuse the AST-walking safe calculator
from facets.tools import Tool, tool

if TYPE_CHECKING:
    from facets.gaia.data import GAIADataset, GAIAQuestion

_MAX_READ_LINES = 200
_MAX_SHEET_ROWS = 500


def build_gaia_tools(dataset: GAIADataset, question: GAIAQuestion) -> list[Tool]:
    """Build the file toolset scoped to one GAIA question's single attachment.

    The attachment is downloaded once (and cached) when the first file tool runs. A question with
    no attachment gets only ``compute`` — the recipe still runs, it just has nothing to read (and
    such a question is, by definition, an open-web task this minimal toolset can't fully serve).
    """
    ext = question.file_ext

    def _local_path() -> str:
        return str(dataset.attachment_path(question))

    @tool(name="describe_attachment")
    async def describe_attachment() -> dict:
        """Report the attached file's name and type, so you know which reader to use."""
        if not question.has_file:
            return {"has_file": False, "note": "This question has no attachment."}
        return {"has_file": True, "file_name": question.file_name, "type": ext}

    @tool(name="read_file")
    async def read_file(start_line: int = 0, max_lines: int = 100) -> dict:
        """Read a window of the attached text file (txt/csv/json/py/…), starting at ``start_line``.

        Page through long files with ``start_line``. Use read_spreadsheet for .xlsx and read_pdf
        for .pdf instead.
        """
        text = _read_text(_local_path())
        lines = text.splitlines()
        start = max(0, int(float(start_line)))
        count = min(max(1, int(float(max_lines))), _MAX_READ_LINES)
        window = lines[start : start + count]
        return {
            "file_name": question.file_name,
            "total_lines": len(lines),
            "start_line": start,
            "lines": {str(start + i): line for i, line in enumerate(window)},
            "more": start + count < len(lines),
        }

    @tool(name="read_spreadsheet")
    async def read_spreadsheet(sheet: str | None = None, max_rows: int = 200) -> dict:
        """Read an .xlsx spreadsheet as rows of cell values. Optionally name a ``sheet``."""
        return _read_xlsx(_local_path(), sheet, min(max(1, int(float(max_rows))), _MAX_SHEET_ROWS))

    @tool(name="read_pdf")
    async def read_pdf(max_pages: int = 20) -> dict:
        """Extract text from the attached PDF, up to ``max_pages`` pages."""
        return _read_pdf(_local_path(), max(1, int(float(max_pages))))

    @tool(name="compute")
    async def compute(expression: str) -> dict:
        """Evaluate an arithmetic expression exactly (+ - * / // % ** and parentheses).

        Use this for arithmetic instead of doing it in your head — GAIA numeric answers are graded
        for exact equality.
        """
        try:
            value = _safe_eval(expression)
        except Exception as exc:  # noqa: BLE001 — surfaced as a soft error to the agent
            return {"expression": expression, "error": f"{type(exc).__name__}: {exc}"}
        return {"expression": expression, "result": value}

    tools: list[Tool] = [compute]
    if question.has_file:
        tools.insert(0, describe_attachment)
        if ext in ("xlsx", "xls"):
            tools.insert(1, read_spreadsheet)
        elif ext == "pdf":
            tools.insert(1, read_pdf)
        else:
            tools.insert(1, read_file)
    return tools


def _read_text(path: str) -> str:
    with open(path, encoding="utf-8", errors="replace") as f:
        return f.read()


def _read_xlsx(path: str, sheet: str | None, max_rows: int) -> dict:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        from facets.config import MissingCredential

        raise MissingCredential(
            "Reading .xlsx attachments needs the 'openpyxl' package.\n"
            "How to fix: `uv add openpyxl` (or `uv run --with openpyxl ...` for a one-off)."
        ) from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    rows: list[list] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        rows.append([("" if c is None else c) for c in row])
    result = {
        "file": path.rsplit("/", 1)[-1],
        "sheet": ws.title,
        "sheet_names": wb.sheetnames,
        "row_count": len(rows),
        "rows": rows,
        "truncated": len(rows) >= max_rows,
    }
    wb.close()
    return result


def _read_pdf(path: str, max_pages: int) -> dict:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        from facets.config import MissingCredential

        raise MissingCredential(
            "Reading .pdf attachments needs the 'pypdf' package.\n"
            "How to fix: `uv add pypdf` (or `uv run --with pypdf ...` for a one-off)."
        ) from exc

    reader = PdfReader(path)
    pages = reader.pages[:max_pages]
    text = "\n\n".join(
        f"--- page {i + 1} ---\n{p.extract_text() or ''}" for i, p in enumerate(pages)
    )
    return {
        "file": path.rsplit("/", 1)[-1],
        "total_pages": len(reader.pages),
        "returned_pages": len(pages),
        "text": text,
    }
